import collections
import csv
import datetime
import itertools
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.conf import settings
from django.core.exceptions import FieldDoesNotExist, ObjectDoesNotExist
from django.db.models import FileField
from django.db.models.fields.related import ManyToManyField, OneToOneField
from django.db.transaction import atomic
from django.http import HttpResponse, StreamingHttpResponse
from django.utils import dateformat
from django.utils.encoding import force_str, smart_str
from django.utils.timezone import get_default_timezone

from adminactions import utils

from .utils import clone_instance, get_field_by_path, get_field_value, get_ignored_fields

csv_options_default = {
    "date_format": "d/m/Y",
    "datetime_format": "N j, Y, P",
    "time_format": "P",
    "header": False,
    "quotechar": '"',
    "quoting": csv.QUOTE_ALL,
    "delimiter": ";",
    "escapechar": "\\",
}

delimiters = ",;|:"
quotes = "'\"`"
escapechars = " \\"
ALL_FIELDS = -999


def merge(master, other, fields=None, commit=False, m2m=None, related=None):  # noqa
    """
        Merge 'other' into master.

        `fields` is a list of fieldnames that must be readed from ``other`` to put into master.
        If ``fields`` is None ``master`` will not get any of the ``other`` values.
        Finally ``other`` will be deleted and master will be preserved

    @param master:  Model instance
    @param other: Model instance
    @param fields: list of fieldnames to  merge
    @param m2m: list of m2m fields to merge. If empty will be removed
    @param related: list of related fieldnames to merge. If empty will be removed
    @return:
    """

    fields = fields or []

    all_m2m = {}
    all_related = {}

    if related == ALL_FIELDS:
        related = [rel.get_accessor_name() for rel in utils.get_all_related_objects(master)]

    if m2m == ALL_FIELDS:
        m2m = set()
        for field in master._meta.get_fields():
            if getattr(field, "many_to_many", None):
                if isinstance(field, ManyToManyField):
                    if not field.remote_field.through._meta.auto_created:
                        continue
                    m2m.add(field.name)
                else:
                    # reverse relation
                    m2m.add(field.get_accessor_name())
    if m2m and not commit:
        raise ValueError("Cannot save related with `commit=False`")
    with atomic():
        result = clone_instance(master)
        for fieldname in fields:
            f = get_field_by_path(master, fieldname)
            if isinstance(f, FileField) or f and not f.primary_key:
                setattr(result, fieldname, getattr(other, fieldname))

        if m2m:
            for accessor in set(m2m):
                all_m2m[accessor] = []
                source_m2m = getattr(other, accessor)
                for r in source_m2m.all():
                    all_m2m[accessor].append(r)
        if related:
            for name in set(related):
                related_object = get_field_by_path(master, name)
                all_related[name] = []
                if related_object and isinstance(related_object.field, OneToOneField):
                    try:
                        accessor = getattr(other, name)
                        all_related[name] = [(related_object.field.name, accessor)]
                    except ObjectDoesNotExist:
                        pass
                else:
                    accessor = getattr(other, name, None)
                    if accessor:
                        rel_fieldname = list(accessor.core_filters.keys())[0].split("__")[0]
                        for r in accessor.all():
                            all_related[name].append((rel_fieldname, r))

        if commit:
            for name, elements in list(all_related.items()):
                for rel_fieldname, element in elements:
                    setattr(element, rel_fieldname, master)
                    element.save()
            other.delete()
            ignored_fields = get_ignored_fields(result._meta.model, "MERGE_ACTION_IGNORED_FIELDS")
            for ig_field in ignored_fields:
                setattr(result, ig_field, result._meta.get_field(ig_field).get_default())
            result.save()
            for fieldname, elements in list(all_m2m.items()):
                dest_m2m = getattr(result, fieldname)
                for element in elements:
                    dest_m2m.add(element)
    return result


class Echo:
    """An object that implements just the write method of the file-like
    interface.
    """

    def write(self, value):
        """Write the value by returning it, instead of storing in a buffer."""
        return value


def export_as_csv(  # noqa: max-complexity: 20
    queryset,
    fields=None,
    header=None,
    filename=None,
    options=None,
    out=None,
    modeladmin=None,
):  # noqa
    """
        Exports a queryset as csv from a queryset with the given fields.

    :param queryset: queryset to export
    :param fields: list of fields names to export. None for all fields
    :param header: if True, the exported file will have the first row as column names
    :param filename: name of the filename
    :param options: CSVOptions() instance or none
    :param: out: object that implements File protocol. HttpResponse if None.

    :return: HttpResponse instance
    """
    streaming_enabled = getattr(settings, "ADMINACTIONS_STREAM_CSV", False)
    if out is None:
        if streaming_enabled:
            response_class = StreamingHttpResponse
        else:
            response_class = HttpResponse

        if filename is None:
            filename = "%s.csv" % queryset.model._meta.verbose_name_plural.lower().replace(" ", "_")

        response = response_class(content_type="text/csv")
        response["Content-Disposition"] = ('attachment;filename="%s"' % filename).encode(
            "us-ascii", "replace"
        )
    else:
        response = out

    if options is None:
        config = csv_options_default
    else:
        config = csv_options_default.copy()
        config.update(options)
    if fields is None:
        fields = [f.name for f in queryset.model._meta.fields + queryset.model._meta.many_to_many]
    if streaming_enabled:
        buffer_object = Echo()
    else:
        buffer_object = response

    dialect = config.get("dialect", None)
    if dialect is not None:
        writer = csv.writer(buffer_object, dialect=dialect)
    else:
        writer = csv.writer(
            buffer_object,
            escapechar=config["escapechar"],
            delimiter=str(config["delimiter"]),
            quotechar=str(config["quotechar"]),
            quoting=int(config["quoting"]),
        )

    settingstime_zone = get_default_timezone()

    def yield_header():
        if bool(header):
            if isinstance(header, (list, tuple)):
                yield writer.writerow(header)
            else:
                yield writer.writerow([f for f in fields])
        yield ""

    def yield_rows():
        for obj in queryset:
            row = []
            for fieldname in fields:
                value = get_field_value(obj, fieldname, modeladmin=modeladmin)
                if isinstance(value, datetime.datetime):
                    try:
                        value = dateformat.format(
                            value.astimezone(settingstime_zone),
                            config["datetime_format"],
                        )
                    except ValueError:
                        # astimezone() cannot be applied to a naive datetime
                        value = dateformat.format(value, config["datetime_format"])
                elif isinstance(value, datetime.date):
                    value = dateformat.format(value, config["date_format"])
                elif isinstance(value, datetime.time):
                    value = dateformat.format(value, config["time_format"])
                row.append(smart_str(value))
            yield writer.writerow(row)

    if streaming_enabled:
        content_attr = "content" if (StreamingHttpResponse is HttpResponse) else "streaming_content"
        setattr(response, content_attr, itertools.chain(yield_header(), yield_rows()))
    else:
        collections.deque(itertools.chain(yield_header(), yield_rows()), maxlen=0)

    return response


xls_options_default = {
    "date_format": "d/m/Y",
    "datetime_format": "N j, Y, P",
    "time_format": "P",
    "sheet_name": "Sheet1",
    "DateField": "DD MMM-YY",
    "DateTimeField": "DD MMD YY hh:mm",
    "TimeField": "hh:mm",
    "IntegerField": "#,##",
    "PositiveIntegerField": "#,##",
    "PositiveSmallIntegerField": "#,##",
    "BigIntegerField": "#,##",
    "DecimalField": "#,##0.00",
    "BooleanField": "boolean",
    "NullBooleanField": "boolean",
    "EmailField": lambda value: 'mailto:%s' % value,
    "URLField": lambda value: str(value),
    "CurrencyColumn": '"$"#,##0.00);[Red]("$"#,##0.00)',
}


def export_as_xls2(  # noqa: max-complexity: 24
    queryset, fields=None, header=None, filename=None, options=None, out=None, modeladmin=None  # noqa
):
    """
    Exports a queryset as xlsx (openpyxl) from a queryset with the given fields.
    Replaces the legacy xlwt-based export. Output format is now .xlsx.

    :param queryset: queryset to export (can also be list of namedtuples)
    :param fields: list of fields names to export. None for all fields
    :param header: if True, the exported file will have the first row as column names
    :param out: object that implements File protocol.
    :param filename: name of the download file
    :param options: dict of options
    :param modeladmin: ModelAdmin instance
    :return: HttpResponse instance if out not supplied, otherwise out
    """

    http_response = out is None
    buffer = BytesIO()

    config = xls_options_default.copy()
    if options:
        config.update(options)

    if fields is None:
        fields = [f.name for f in queryset.model._meta.fields + queryset.model._meta.many_to_many]

    sheet_name = config.pop("sheet_name", "Sheet1")
    use_display = config.get("use_display", False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    heading_font = Font(bold=True, size=11)
    heading_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = 1
    # Write row number header cell
    ws.cell(row=row_num, column=1, value="#").font = heading_font

    if header:
        if not isinstance(header, (list, tuple)):
            header = [
                force_str(f.verbose_name)
                for f in queryset.model._meta.fields + queryset.model._meta.many_to_many
                if f.name in fields
            ]
        for col_idx, fieldname in enumerate(header, start=2):
            cell = ws.cell(row=row_num, column=col_idx, value=fieldname)
            cell.font = heading_font
            cell.alignment = heading_align
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    settingstime_zone = get_default_timezone()

    for rownum, obj in enumerate(queryset, start=1):
        ws.cell(row=rownum + 1, column=1, value=rownum)
        for col_idx, fieldname in enumerate(fields, start=2):
            try:
                value = get_field_value(
                    obj, fieldname, usedisplay=use_display, raw_callable=False, modeladmin=modeladmin
                )
                # Format dates/times as strings
                if isinstance(value, datetime.datetime):
                    try:
                        value = dateformat.format(
                            value.astimezone(settingstime_zone),
                            config.get("datetime_format", "N j, Y, P"),
                        )
                    except ValueError:
                        value = dateformat.format(value, config.get("datetime_format", "N j, Y, P"))
                elif isinstance(value, datetime.date):
                    value = dateformat.format(value, config.get("date_format", "d/m/Y"))
                elif isinstance(value, datetime.time):
                    value = dateformat.format(value, config.get("time_format", "P"))
                elif callable(value):
                    value = value()
                if isinstance(value, (list, tuple)):
                    value = ", ".join(str(v) for v in value)
                ws.cell(row=rownum + 1, column=col_idx, value=smart_str(value) if value is not None else "")
            except Exception as e:
                ws.cell(row=rownum + 1, column=col_idx, value=smart_str(e))

    wb.save(buffer)
    buffer.seek(0)

    if http_response:
        if filename is None:
            filename = "%s.xlsx" % queryset.model._meta.verbose_name_plural.lower().replace(" ", "_")
        elif filename.endswith(".xls"):
            filename = filename + "x"  # .xls -> .xlsx
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment;filename="%s"' % filename
        return response
    return buffer


export_as_xls = export_as_xls2
